"""
TokoKu POS - Backend Server (Extended v2.0)
Flask + SQLite | Modul: Distributor, Pembelian, Hutang, Kas, Laporan Keuangan
"""

from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import sqlite3, hashlib, json, os, datetime, traceback, io

app = Flask(__name__, static_folder='static')
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, 'tokoku.db')

# ─── DATABASE ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    # Tabel lama (kompatibel)
    c.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id       INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role     TEXT NOT NULL DEFAULT 'cashier',
        name     TEXT NOT NULL,
        active   INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS categories (
        id   INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL
    );
    CREATE TABLE IF NOT EXISTS products (
        id                     INTEGER PRIMARY KEY AUTOINCREMENT,
        barcode                TEXT,
        name                   TEXT NOT NULL,
        category               TEXT,
        buy_price              REAL DEFAULT 0,
        price                  REAL NOT NULL,
        wholesale_price        REAL DEFAULT NULL,
        wholesale_price_manual INTEGER DEFAULT 0,
        stock                  REAL DEFAULT 0,
        min_stock              REAL DEFAULT 5,
        unit                   TEXT DEFAULT 'pcs',
        active                 INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS customers (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        name         TEXT NOT NULL,
        phone        TEXT,
        address      TEXT,
        notes        TEXT,
        pricing_type TEXT DEFAULT 'retail',
        active       INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS transactions (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        date           TEXT NOT NULL,
        cashier_id     INTEGER,
        cashier_name   TEXT,
        subtotal       REAL DEFAULT 0,
        discount       REAL DEFAULT 0,
        tax            REAL DEFAULT 0,
        total          REAL DEFAULT 0,
        payment_method TEXT DEFAULT 'cash',
        payment_amount REAL DEFAULT 0,
        customer_id    INTEGER,
        customer_name  TEXT,
        status         TEXT DEFAULT 'done',
        void_reason    TEXT,
        voided_by      TEXT,
        voided_at      TEXT
    );
    CREATE TABLE IF NOT EXISTS transaction_items (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        transaction_id INTEGER NOT NULL,
        product_id     INTEGER,
        name           TEXT,
        qty            REAL DEFAULT 1,
        price          REAL DEFAULT 0,
        buy_price      REAL DEFAULT 0,
        subtotal       REAL DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS debts (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_id    INTEGER NOT NULL,
        transaction_id INTEGER,
        amount         REAL DEFAULT 0,
        paid           REAL DEFAULT 0,
        date           TEXT,
        note           TEXT
    );
    CREATE TABLE IF NOT EXISTS debt_payments (
        id      INTEGER PRIMARY KEY AUTOINCREMENT,
        debt_id INTEGER NOT NULL,
        amount  REAL DEFAULT 0,
        date    TEXT,
        note    TEXT
    );
    CREATE TABLE IF NOT EXISTS stock_log (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id   INTEGER,
        product_name TEXT,
        delta        REAL,
        reason       TEXT,
        before       REAL,
        after        REAL,
        user_id      INTEGER,
        date         TEXT
    );
    CREATE TABLE IF NOT EXISTS activity_log (
        id      INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        action  TEXT,
        detail  TEXT,
        date    TEXT
    );
    CREATE TABLE IF NOT EXISTS settings (
        key   TEXT PRIMARY KEY,
        value TEXT
    );
    """)

    # ── TABEL BARU ──
    c.executescript("""
    CREATE TABLE IF NOT EXISTS distributors (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        code       TEXT UNIQUE,
        name       TEXT NOT NULL,
        sales_name TEXT,
        phone      TEXT,
        address    TEXT,
        notes      TEXT,
        active     INTEGER DEFAULT 1,
        created_at TEXT,
        updated_at TEXT
    );

    CREATE TABLE IF NOT EXISTS purchase_invoices (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_no      TEXT UNIQUE NOT NULL,
        invoice_date    TEXT NOT NULL,
        due_date        TEXT,
        distributor_id  INTEGER NOT NULL,
        distributor_name TEXT,
        payment_type    TEXT NOT NULL,
        total_amount    REAL DEFAULT 0,
        paid_amount     REAL DEFAULT 0,
        remaining_amount REAL DEFAULT 0,
        status          TEXT DEFAULT 'unpaid',
        notes           TEXT,
        created_by      INTEGER,
        created_by_name TEXT,
        created_at      TEXT,
        updated_by      INTEGER,
        updated_at      TEXT,
        FOREIGN KEY (distributor_id) REFERENCES distributors(id)
    );

    CREATE TABLE IF NOT EXISTS purchase_invoice_items (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_id  INTEGER NOT NULL,
        product_id  INTEGER,
        product_name TEXT,
        qty         REAL DEFAULT 0,
        buy_price   REAL DEFAULT 0,
        subtotal    REAL DEFAULT 0,
        FOREIGN KEY (invoice_id) REFERENCES purchase_invoices(id)
    );

    CREATE TABLE IF NOT EXISTS supplier_debt_payments (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_id   INTEGER NOT NULL,
        payment_date TEXT NOT NULL,
        amount       REAL DEFAULT 0,
        method       TEXT DEFAULT 'transfer',
        notes        TEXT,
        created_by   INTEGER,
        created_by_name TEXT,
        created_at   TEXT,
        FOREIGN KEY (invoice_id) REFERENCES purchase_invoices(id)
    );

    CREATE TABLE IF NOT EXISTS cash_transactions (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        type        TEXT NOT NULL,
        date        TEXT NOT NULL,
        category    TEXT,
        reference   TEXT,
        description TEXT,
        amount      REAL DEFAULT 0,
        balance     REAL DEFAULT 0,
        source      TEXT,
        source_id   INTEGER,
        created_by  INTEGER,
        created_by_name TEXT,
        created_at  TEXT
    );
    """)

    conn.commit()

    # ── MIGRASI KOLOM BARU (kompatibel dengan DB lama) ──
    existing_prod = [r[1] for r in c.execute('PRAGMA table_info(products)').fetchall()]
    existing_cust = [r[1] for r in c.execute('PRAGMA table_info(customers)').fetchall()]
    if 'wholesale_price' not in existing_prod:
        c.execute('ALTER TABLE products ADD COLUMN wholesale_price REAL DEFAULT NULL')
    if 'wholesale_price_manual' not in existing_prod:
        c.execute('ALTER TABLE products ADD COLUMN wholesale_price_manual INTEGER DEFAULT 0')
    if 'pricing_type' not in existing_cust:
        c.execute('ALTER TABLE customers ADD COLUMN pricing_type TEXT DEFAULT "retail"')
    # Set default wholesale = 90% harga eceran untuk produk lama yg belum punya grosir
    c.execute('''UPDATE products SET wholesale_price=ROUND(price*0.9,0), wholesale_price_manual=0
                 WHERE wholesale_price IS NULL AND price > 0''')
    conn.commit()
    if c.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        def h(p): return hashlib.sha256(p.encode()).hexdigest()
        c.executemany("INSERT INTO users (username,password,role,name) VALUES (?,?,?,?)", [
            ('owner', h('owner123'), 'owner', 'Bapak Haji Sugiono'),
            ('admin', h('admin123'), 'admin', 'Admin Toko'),
            ('kasir', h('kasir123'), 'cashier', 'Kasir Satu'),
        ])
        cats = ['Mie Instan','Minuman','Snack','Sembako','Rokok','Kebersihan','Lainnya']
        c.executemany("INSERT INTO categories (name) VALUES (?)", [(x,) for x in cats])
        products = [
            ('8992388001','Indomie Goreng','Mie Instan',2800,3500,None,0,150,30,'pcs'),
            ('8992388002','Indomie Kuah Ayam','Mie Instan',2800,3500,None,0,120,30,'pcs'),
            ('8992388003','Indomie Soto','Mie Instan',2800,3500,None,0,95,30,'pcs'),
            ('8998889001','Aqua 600ml','Minuman',2500,3000,None,0,200,50,'pcs'),
            ('8998889002','Aqua 1500ml','Minuman',4500,5500,None,0,80,24,'pcs'),
            ('8887290001','Teh Botol Sosro 350ml','Minuman',3800,5000,None,0,72,24,'pcs'),
            ('8992761001','Chitato Sapi Panggang 68g','Snack',6500,8000,None,0,48,12,'pcs'),
            ('8992761002','Lays Original 68g','Snack',7000,9000,None,0,36,12,'pcs'),
            ('8991100001','Roma Kelapa 300g','Snack',11000,14000,None,0,30,8,'pcs'),
            ('8888888001','Beras Premium 5kg','Sembako',68000,78000,None,0,25,5,'karung'),
            ('8888888002','Gula Pasir 1kg','Sembako',14000,16000,None,0,40,10,'kg'),
            ('8888888003','Minyak Goreng Bimoli 2L','Sembako',28000,32000,None,0,30,10,'botol'),
            ('8998900002','Dancow Full Cream 400g','Sembako',32000,38000,None,0,5,6,'pcs'),
            ('8998776001','Sabun Lifebuoy 80g','Kebersihan',4500,6000,None,0,60,15,'pcs'),
            ('8998776002','Rinso Anti Noda 1kg','Kebersihan',18000,22000,None,0,24,6,'pcs'),
            ('8887300001','Sunlight Jeruk 400ml','Kebersihan',8500,11000,None,0,18,6,'pcs'),
            ('8886780001','Djarum Super 12','Rokok',23000,26000,None,0,100,20,'pack'),
            ('8886780002','Gudang Garam 12','Rokok',24000,27000,None,0,80,20,'pack'),
            ('8992900001','Energen Coklat 300g','Minuman',10000,13000,None,0,36,10,'pcs'),
            ('8881200001','Kopi Kapal Api 165g','Minuman',14000,18000,None,0,0,5,'pcs'),
        ]
        c.executemany("INSERT INTO products (barcode,name,category,buy_price,price,wholesale_price,wholesale_price_manual,stock,min_stock,unit) VALUES (?,?,?,?,?,?,?,?,?,?)", products)
        # Set default wholesale 10% untuk seed data
        c.execute("UPDATE products SET wholesale_price=ROUND(price*0.9,0) WHERE wholesale_price IS NULL")
        customers = [
            ('Bu Ratna Wijaya','08121234567','Jl. Mawar No.5','Pelanggan tetap'),
            ('Pak Budi Santoso','08987654321','RT 03 RW 02',''),
            ('Bu Sari Dewi','08567890123','Gang Kenanga No.12','Sering hutang'),
            ('Pak Hendra','08234567890','',''),
            ('Bu Dewi Rahayu','08345678901','Komplek Griya Asri','Bayar bulanan'),
        ]
        c.executemany("INSERT INTO customers (name,phone,address,notes) VALUES (?,?,?,?)", customers)
        c.executemany("INSERT INTO settings (key,value) VALUES (?,?)", [
            ('storeName','Toko Kelontong Maju'),
            ('storeAddr','Jl. Merdeka No. 17'),
            ('tax','0'), ('taxPct','11'), ('lang','id'),
            ('receiptFooter','Terima kasih sudah berbelanja!'),
            ('bankInfo','BCA: 1234-5678-90 a.n. Toko Kelontong Maju'),
        ])
        today = datetime.date.today().isoformat()
        c.execute("INSERT INTO debts (customer_id,amount,paid,date,note) VALUES (1,35000,0,?,'Belanja sembako')", (today,))
        c.execute("INSERT INTO debts (customer_id,amount,paid,date,note) VALUES (3,76000,30000,?,'Belanja bulanan')", (today,))
        c.execute("INSERT INTO debts (customer_id,amount,paid,date,note) VALUES (5,125000,0,?,'Belanja akhir bulan')", (today,))

        # Seed distributor demo
        now = datetime.datetime.now().isoformat()
        c.execute("""INSERT INTO distributors (code,name,sales_name,phone,address,notes,created_at,updated_at)
                     VALUES (?,?,?,?,?,?,?,?)""",
                  ('DIST001','PT Indo Sejahtera','Pak Agus','081234567890','Jl. Industri No.5 Jakarta','Distributor mie & snack',now,now))
        c.execute("""INSERT INTO distributors (code,name,sales_name,phone,address,notes,created_at,updated_at)
                     VALUES (?,?,?,?,?,?,?,?)""",
                  ('DIST002','CV Maju Bersama','Bu Rina','082345678901','Jl. Raya Bekasi No.12','Distributor minuman & dairy',now,now))
        c.execute("""INSERT INTO distributors (code,name,sales_name,phone,address,notes,created_at,updated_at)
                     VALUES (?,?,?,?,?,?,?,?)""",
                  ('DIST003','UD Sembako Jaya','Pak Hadi','083456789012','Pasar Induk Cibitung','Distributor sembako',now,now))
        conn.commit()
    conn.close()

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def ok(data=None, **kw):
    res = {'ok': True}
    if data is not None: res['data'] = data
    res.update(kw)
    return jsonify(res)

def err(msg, code=400):
    return jsonify({'ok': False, 'error': msg}), code

def row2dict(row):
    return dict(row) if row else None

def rows2list(rows):
    return [dict(r) for r in rows]

def now_iso():
    return datetime.datetime.now().isoformat()

def today_str():
    return datetime.date.today().isoformat()

def log_activity(user_id, action, detail):
    conn = get_db()
    conn.execute("INSERT INTO activity_log (user_id,action,detail,date) VALUES (?,?,?,?)",
                 (user_id, action, detail, now_iso()))
    conn.commit()
    conn.close()

def get_cash_balance(conn=None):
    """Hitung saldo kas dari semua transaksi kas."""
    close_after = conn is None
    if conn is None:
        conn = get_db()
    masuk = conn.execute("SELECT COALESCE(SUM(amount),0) FROM cash_transactions WHERE type='in'").fetchone()[0]
    keluar = conn.execute("SELECT COALESCE(SUM(amount),0) FROM cash_transactions WHERE type='out'").fetchone()[0]
    if close_after:
        conn.close()
    return masuk - keluar

def update_invoice_status(conn, invoice_id):
    """Update status faktur berdasarkan remaining_amount dan due_date."""
    inv = row2dict(conn.execute("SELECT * FROM purchase_invoices WHERE id=?", (invoice_id,)).fetchone())
    if not inv:
        return
    total_paid = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM supplier_debt_payments WHERE invoice_id=?",
        (invoice_id,)
    ).fetchone()[0]
    # Untuk cash: paid_amount = total, remaining = 0
    if inv['payment_type'] == 'cash':
        remaining = 0
        status = 'paid'
    else:
        remaining = inv['total_amount'] - total_paid
        if remaining <= 0:
            remaining = 0
            status = 'paid'
        else:
            today = today_str()
            due = inv['due_date'] or ''
            if due and due < today:
                status = 'overdue'
            elif total_paid > 0:
                status = 'partial'
            else:
                status = 'unpaid'
    conn.execute("""UPDATE purchase_invoices
                    SET paid_amount=?, remaining_amount=?, status=?, updated_at=?
                    WHERE id=?""",
                 (total_paid, remaining, status, now_iso(), invoice_id))

def recalculate_all_invoice_statuses():
    """Recalculate overdue status untuk semua faktur aktif."""
    conn = get_db()
    invoices = rows2list(conn.execute(
        "SELECT id FROM purchase_invoices WHERE status IN ('unpaid','partial')"
    ).fetchall())
    for inv in invoices:
        update_invoice_status(conn, inv['id'])
    conn.commit()
    conn.close()

def add_cash_transaction(conn, type_, category, reference, description, amount, source, source_id, user_id, user_name, date=None):
    """Tambah transaksi kas dan update saldo."""
    balance = get_cash_balance(conn)
    if type_ == 'in':
        new_balance = balance + amount
    else:
        new_balance = balance - amount
    conn.execute("""INSERT INTO cash_transactions
                    (type,date,category,reference,description,amount,balance,source,source_id,created_by,created_by_name,created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                 (type_, date or today_str(), category, reference, description, amount,
                  new_balance, source, source_id, user_id, user_name, now_iso()))
    return new_balance

# ─── AUTH ─────────────────────────────────────────────────────────────────────

@app.route('/api/login', methods=['POST'])
def login():
    d = request.json or {}
    username = d.get('username','').strip()
    password = d.get('password','')
    if not username or not password:
        return err('Username dan password wajib diisi')
    hashed = hashlib.sha256(password.encode()).hexdigest()
    conn = get_db()
    user = conn.execute(
        "SELECT id,username,role,name FROM users WHERE username=? AND password=? AND active=1",
        (username, hashed)
    ).fetchone()
    conn.close()
    if not user:
        return err('Username atau password salah', 401)
    u = dict(user)
    log_activity(u['id'], 'login', u['name'] + ' login')
    return ok(u)

@app.route('/api/logout', methods=['POST'])
def logout():
    d = request.json or {}
    log_activity(d.get('user_id'), 'logout', d.get('name','') + ' logout')
    return ok()

# ─── SETTINGS ────────────────────────────────────────────────────────────────

@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = get_db()
    rows = conn.execute("SELECT key,value FROM settings").fetchall()
    conn.close()
    return ok({r['key']: r['value'] for r in rows})

@app.route('/api/settings', methods=['PUT'])
def put_settings():
    d = request.json or {}
    conn = get_db()
    for k, v in d.items():
        conn.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (k, str(v)))
    conn.commit()
    conn.close()
    return ok()

# ─── CATEGORIES ──────────────────────────────────────────────────────────────

@app.route('/api/categories', methods=['GET'])
def get_categories():
    conn = get_db()
    rows = conn.execute("SELECT * FROM categories ORDER BY name").fetchall()
    conn.close()
    return ok(rows2list(rows))

@app.route('/api/categories', methods=['POST'])
def add_category():
    d = request.json or {}
    name = d.get('name','').strip()
    if not name: return err('Nama kategori wajib diisi')
    conn = get_db()
    try:
        c = conn.execute("INSERT INTO categories (name) VALUES (?)", (name,))
        conn.commit()
        cat_id = c.lastrowid
    except sqlite3.IntegrityError:
        conn.close()
        return err('Kategori sudah ada')
    conn.close()
    return ok({'id': cat_id, 'name': name})

@app.route('/api/categories/<int:cat_id>', methods=['DELETE'])
def del_category(cat_id):
    conn = get_db()
    conn.execute("DELETE FROM categories WHERE id=?", (cat_id,))
    conn.commit()
    conn.close()
    return ok()

# ─── PRODUCTS ────────────────────────────────────────────────────────────────

@app.route('/api/products', methods=['GET'])
def get_products():
    conn = get_db()
    rows = conn.execute("SELECT * FROM products WHERE active=1 ORDER BY name").fetchall()
    conn.close()
    return ok(rows2list(rows))

@app.route('/api/products', methods=['POST'])
def add_product():
    d = request.json or {}
    if not d.get('name') or not d.get('price'):
        return err('Nama dan harga wajib diisi')
    retail = float(d['price'])
    # wholesale: pakai nilai dari form jika diisi manual, else None (akan dihitung otomatis)
    ws_raw = d.get('wholesale_price')
    ws_manual = 1 if (ws_raw is not None and str(ws_raw).strip() != '') else 0
    wholesale = float(ws_raw) if ws_manual else round(retail * 0.9, 0)
    conn = get_db()
    c = conn.execute("""
        INSERT INTO products (barcode,name,category,buy_price,price,wholesale_price,wholesale_price_manual,stock,min_stock,unit)
        VALUES (:barcode,:name,:category,:buy_price,:price,:wholesale_price,:wholesale_price_manual,:stock,:min_stock,:unit)
    """, {
        'barcode':                d.get('barcode',''),
        'name':                   d['name'].strip(),
        'category':               d.get('category','Lainnya'),
        'buy_price':              float(d.get('buy_price') or d.get('buyPrice') or 0),
        'price':                  retail,
        'wholesale_price':        wholesale,
        'wholesale_price_manual': ws_manual,
        'stock':                  float(d.get('stock',0)),
        'min_stock':              float(d.get('min_stock') or d.get('minStock') or 5),
        'unit':                   d.get('unit','pcs'),
    })
    conn.commit()
    new_id = c.lastrowid
    prod = row2dict(conn.execute("SELECT * FROM products WHERE id=?", (new_id,)).fetchone())
    conn.close()
    log_activity(d.get('user_id'), 'add_product', 'Tambah: '+d['name'])
    return ok(prod)

@app.route('/api/products/<int:pid>', methods=['PUT'])
def update_product(pid):
    d = request.json or {}
    retail = float(d.get('price', 0))
    # Logika wholesale:
    # - Jika wholesale_price dikirim dan tidak kosong → manual, simpan nilai tsb
    # - Jika wholesale_price_manual=0 dikirim eksplisit → reset ke otomatis (90% retail)
    # - Jika harga eceran berubah dan manual=0 → recalculate otomatis
    ws_raw    = d.get('wholesale_price')
    ws_manual = d.get('wholesale_price_manual')
    if ws_manual == 0:
        # Kasir/admin reset ke otomatis
        wholesale = round(retail * 0.9, 0)
        ws_manual_val = 0
    elif ws_raw is not None and str(ws_raw).strip() != '':
        wholesale = float(ws_raw)
        ws_manual_val = 1
    else:
        # Tidak ada perubahan wholesale → baca dari DB, recalculate jika otomatis
        conn2 = get_db()
        existing = row2dict(conn2.execute("SELECT wholesale_price, wholesale_price_manual FROM products WHERE id=?", (pid,)).fetchone())
        conn2.close()
        if existing and existing['wholesale_price_manual'] == 0:
            wholesale = round(retail * 0.9, 0)
            ws_manual_val = 0
        else:
            wholesale = existing['wholesale_price'] if existing else round(retail * 0.9, 0)
            ws_manual_val = 1

    conn = get_db()
    conn.execute("""
        UPDATE products SET barcode=:barcode,name=:name,category=:category,
        buy_price=:buy_price,price=:price,wholesale_price=:wholesale_price,
        wholesale_price_manual=:wholesale_price_manual,
        stock=:stock,min_stock=:min_stock,unit=:unit
        WHERE id=:id
    """, {
        'id':                     pid,
        'barcode':                d.get('barcode',''),
        'name':                   d.get('name','').strip(),
        'category':               d.get('category','Lainnya'),
        'buy_price':              float(d.get('buy_price') or d.get('buyPrice') or 0),
        'price':                  retail,
        'wholesale_price':        wholesale,
        'wholesale_price_manual': ws_manual_val,
        'stock':                  float(d.get('stock',0)),
        'min_stock':              float(d.get('min_stock') or d.get('minStock') or 5),
        'unit':                   d.get('unit','pcs'),
    })
    conn.commit()
    prod = row2dict(conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone())
    conn.close()
    log_activity(d.get('user_id'), 'edit_product', 'Edit: '+d.get('name',''))
    return ok(prod)

@app.route('/api/products/<int:pid>', methods=['DELETE'])
def delete_product(pid):
    d = request.json or {}
    conn = get_db()
    name = row2dict(conn.execute("SELECT name FROM products WHERE id=?", (pid,)).fetchone()) or {}
    conn.execute("UPDATE products SET active=0 WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    log_activity(d.get('user_id'), 'del_product', 'Hapus: '+name.get('name',''))
    return ok()

@app.route('/api/products/<int:pid>/stock', methods=['POST'])
def adjust_stock(pid):
    d = request.json or {}
    delta = float(d.get('delta',0))
    reason = d.get('reason','')
    if not reason: return err('Alasan wajib diisi')
    conn = get_db()
    prod = row2dict(conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone())
    if not prod: conn.close(); return err('Produk tidak ditemukan', 404)
    before = prod['stock']
    after = before + delta
    if after < 0: conn.close(); return err('Stok tidak bisa negatif')
    conn.execute("UPDATE products SET stock=? WHERE id=?", (after, pid))
    conn.execute("INSERT INTO stock_log (product_id,product_name,delta,reason,before,after,user_id,date) VALUES (?,?,?,?,?,?,?,?)",
                 (pid, prod['name'], delta, reason, before, after, d.get('user_id'), now_iso()))
    conn.commit()
    conn.close()
    log_activity(d.get('user_id'), 'stock_adj', f"Stok {prod['name']}: {delta:+.0f} ({reason})")
    return ok({'stock': after})

# ─── CUSTOMERS ───────────────────────────────────────────────────────────────

@app.route('/api/customers', methods=['GET'])
def get_customers():
    conn = get_db()
    custs = rows2list(conn.execute("SELECT * FROM customers WHERE active=1 ORDER BY name").fetchall())
    debts = rows2list(conn.execute("SELECT customer_id,SUM(amount) as total,SUM(paid) as paid FROM debts GROUP BY customer_id").fetchall())
    conn.close()
    debt_map = {d['customer_id']: d for d in debts}
    for c in custs:
        dm = debt_map.get(c['id'], {})
        c['debt_total'] = dm.get('total', 0) or 0
        c['debt_paid']  = dm.get('paid', 0) or 0
        c['debt_rem']   = c['debt_total'] - c['debt_paid']
    return ok(custs)

@app.route('/api/customers', methods=['POST'])
def add_customer():
    d = request.json or {}
    if not d.get('name'): return err('Nama wajib diisi')
    pricing_type = d.get('pricing_type', 'retail')
    if pricing_type not in ('retail', 'wholesale'): pricing_type = 'retail'
    conn = get_db()
    c = conn.execute("INSERT INTO customers (name,phone,address,notes,pricing_type) VALUES (?,?,?,?,?)",
                     (d['name'].strip(), d.get('phone',''), d.get('address',''), d.get('notes',''), pricing_type))
    conn.commit()
    cust = row2dict(conn.execute("SELECT * FROM customers WHERE id=?", (c.lastrowid,)).fetchone())
    conn.close()
    return ok(cust)

@app.route('/api/customers/<int:cid>', methods=['PUT'])
def update_customer(cid):
    d = request.json or {}
    pricing_type = d.get('pricing_type', 'retail')
    if pricing_type not in ('retail', 'wholesale'): pricing_type = 'retail'
    conn = get_db()
    conn.execute("UPDATE customers SET name=?,phone=?,address=?,notes=?,pricing_type=? WHERE id=?",
                 (d.get('name',''), d.get('phone',''), d.get('address',''), d.get('notes',''), pricing_type, cid))
    conn.commit()
    conn.close()
    return ok()

@app.route('/api/customers/<int:cid>/debts', methods=['GET'])
def get_customer_debts(cid):
    conn = get_db()
    debts = rows2list(conn.execute("SELECT * FROM debts WHERE customer_id=? ORDER BY date DESC", (cid,)).fetchall())
    conn.close()
    return ok(debts)

@app.route('/api/debts/pay', methods=['POST'])
def pay_debt():
    d = request.json or {}
    cid    = d.get('customer_id')
    amount = float(d.get('amount', 0))
    if not cid or amount <= 0: return err('Data tidak lengkap')
    conn = get_db()
    debts = conn.execute("SELECT * FROM debts WHERE customer_id=? AND amount > paid ORDER BY date ASC", (cid,)).fetchall()
    rem = amount
    for debt in debts:
        if rem <= 0: break
        unpaid = debt['amount'] - debt['paid']
        pay = min(rem, unpaid)
        conn.execute("UPDATE debts SET paid=paid+? WHERE id=?", (pay, debt['id']))
        conn.execute("INSERT INTO debt_payments (debt_id,amount,date,note) VALUES (?,?,?,?)",
                     (debt['id'], pay, today_str(), 'Bayar hutang'))
        rem -= pay
    conn.commit()
    conn.close()
    log_activity(d.get('user_id'), 'pay_debt', f"Bayar hutang customer #{cid}: Rp {amount:,.0f}")
    return ok()

# ─── TRANSACTIONS (KASIR) ────────────────────────────────────────────────────

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    limit  = int(request.args.get('limit', 200))
    offset = int(request.args.get('offset', 0))
    date_from = request.args.get('from')
    date_to   = request.args.get('to')
    conn = get_db()
    q = "SELECT * FROM transactions WHERE 1=1"
    params = []
    if date_from: q += " AND date >= ?"; params.append(date_from)
    if date_to:   q += " AND date <= ?"; params.append(date_to + 'T23:59:59')
    q += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params += [limit, offset]
    txs = rows2list(conn.execute(q, params).fetchall())
    conn.close()
    return ok(txs)

@app.route('/api/transactions', methods=['POST'])
def add_transaction():
    d = request.json or {}
    items = d.get('items', [])
    if not items: return err('Tidak ada item')
    conn = get_db()
    try:
        for item in items:
            pid = item.get('id') or item.get('product_id')
            qty = float(item.get('qty', 1))
            prod = row2dict(conn.execute("SELECT * FROM products WHERE id=? AND active=1", (pid,)).fetchone())
            if not prod: conn.close(); return err(f"Produk #{pid} tidak ditemukan")
            if prod['stock'] < qty: conn.close(); return err(f"Stok {prod['name']} tidak cukup (tersisa {prod['stock']})")

        subtotal = float(d.get('subtotal', 0))
        discount = float(d.get('discount', 0))
        tax      = float(d.get('tax', 0))
        total    = float(d.get('total', 0))
        tx_date  = d.get('date', now_iso())

        c = conn.execute("""
            INSERT INTO transactions
            (date,cashier_id,cashier_name,subtotal,discount,tax,total,payment_method,payment_amount,customer_id,customer_name,status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            tx_date, d.get('cashier_id'), d.get('cashier_name',''),
            subtotal, discount, tax, total,
            d.get('payment_method','cash'),
            float(d.get('payment_amount', total)),
            d.get('customer_id'), d.get('customer_name'),
            d.get('status','done')
        ))
        tx_id = c.lastrowid

        for item in items:
            pid   = item.get('id') or item.get('product_id')
            qty   = float(item.get('qty', 1))
            price = float(item.get('price', 0))
            bprice= float(item.get('buyPrice') or item.get('buy_price') or 0)
            conn.execute("""
                INSERT INTO transaction_items (transaction_id,product_id,name,qty,price,buy_price,subtotal)
                VALUES (?,?,?,?,?,?,?)
            """, (tx_id, pid, item.get('name',''), qty, price, bprice, price * qty))
            conn.execute("UPDATE products SET stock = stock - ? WHERE id=?", (qty, pid))

        if d.get('status') in ('partial', 'hutang') and d.get('customer_id'):
            paid_amt = float(d.get('payment_amount', 0))
            rem = total - paid_amt
            if rem > 0:
                conn.execute("INSERT INTO debts (customer_id,transaction_id,amount,paid,date,note) VALUES (?,?,?,?,?,?)",
                             (d['customer_id'], tx_id, rem, 0, today_str(), f"Tx #{tx_id}"))

        # Kas masuk otomatis dari penjualan tunai
        if d.get('payment_method') in ('cash','tunai','qris','transfer') and d.get('status') == 'done':
            user_name = d.get('cashier_name','Kasir')
            add_cash_transaction(conn, 'in', 'Penjualan', f"TRX-{tx_id}",
                                 f"Penjualan - {d.get('cashier_name','')}",
                                 total, 'transaction', tx_id,
                                 d.get('cashier_id'), user_name, tx_date[:10])

        conn.commit()
        tx = row2dict(conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone())
        conn.close()
        log_activity(d.get('cashier_id'), 'transaction', f"Tx #{tx_id} Rp {total:,.0f} via {d.get('payment_method','cash')}")
        return ok(tx)

    except Exception as e:
        conn.close()
        traceback.print_exc()
        return err(str(e), 500)

@app.route('/api/transactions/<int:tx_id>/items', methods=['GET'])
def get_tx_items(tx_id):
    conn = get_db()
    items = rows2list(conn.execute("SELECT * FROM transaction_items WHERE transaction_id=?", (tx_id,)).fetchall())
    conn.close()
    return ok(items)

@app.route('/api/transactions/<int:tx_id>/void', methods=['POST'])
def void_transaction(tx_id):
    d = request.json or {}
    reason = d.get('reason','')
    if not reason: return err('Alasan pembatalan wajib diisi')
    conn = get_db()
    tx = row2dict(conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone())
    if not tx: conn.close(); return err('Transaksi tidak ditemukan', 404)
    if tx['status'] == 'void': conn.close(); return err('Transaksi sudah dibatalkan')
    conn.execute("UPDATE transactions SET status='void',void_reason=?,voided_by=?,voided_at=? WHERE id=?",
                 (reason, d.get('voided_by',''), now_iso(), tx_id))
    conn.commit()
    conn.close()
    log_activity(d.get('user_id'), 'void_tx', f"Batal Tx #{tx_id}: {reason}")
    return ok()

# ─── DISTRIBUTOR ─────────────────────────────────────────────────────────────

@app.route('/api/distributors', methods=['GET'])
def get_distributors():
    recalculate_all_invoice_statuses()
    conn = get_db()
    dists = rows2list(conn.execute("SELECT * FROM distributors ORDER BY name").fetchall())
    for d in dists:
        stats = row2dict(conn.execute("""
            SELECT COUNT(*) as total_invoices,
                   COALESCE(SUM(total_amount),0) as total_purchase,
                   COALESCE(SUM(remaining_amount),0) as total_debt,
                   COALESCE(SUM(paid_amount),0) as total_paid,
                   COUNT(CASE WHEN status='overdue' THEN 1 END) as overdue_count,
                   COUNT(CASE WHEN status IN ('unpaid','partial','overdue') THEN 1 END) as active_invoices
            FROM purchase_invoices WHERE distributor_id=?
        """, (d['id'],)).fetchone())
        d.update(stats or {})
    conn.close()
    return ok(dists)

@app.route('/api/distributors', methods=['POST'])
def add_distributor():
    d = request.json or {}
    if not d.get('name'): return err('Nama distributor wajib diisi')
    conn = get_db()
    # Auto generate code
    last = conn.execute("SELECT code FROM distributors ORDER BY id DESC LIMIT 1").fetchone()
    if last and last['code']:
        try:
            num = int(last['code'].replace('DIST','')) + 1
        except:
            num = conn.execute("SELECT COUNT(*)+1 FROM distributors").fetchone()[0]
    else:
        num = 1
    code = f"DIST{num:03d}"
    try:
        c = conn.execute("""INSERT INTO distributors (code,name,sales_name,phone,address,notes,active,created_at,updated_at)
                           VALUES (?,?,?,?,?,?,1,?,?)""",
                        (code, d['name'].strip(), d.get('sales_name',''), d.get('phone',''),
                         d.get('address',''), d.get('notes',''), now_iso(), now_iso()))
        conn.commit()
        dist = row2dict(conn.execute("SELECT * FROM distributors WHERE id=?", (c.lastrowid,)).fetchone())
    except sqlite3.IntegrityError as e:
        conn.close()
        return err(str(e))
    conn.close()
    return ok(dist)

@app.route('/api/distributors/<int:did>', methods=['PUT'])
def update_distributor(did):
    d = request.json or {}
    conn = get_db()
    conn.execute("""UPDATE distributors SET name=?,sales_name=?,phone=?,address=?,notes=?,active=?,updated_at=?
                    WHERE id=?""",
                 (d.get('name',''), d.get('sales_name',''), d.get('phone',''),
                  d.get('address',''), d.get('notes',''), d.get('active',1), now_iso(), did))
    conn.commit()
    dist = row2dict(conn.execute("SELECT * FROM distributors WHERE id=?", (did,)).fetchone())
    conn.close()
    return ok(dist)

@app.route('/api/distributors/<int:did>', methods=['DELETE'])
def delete_distributor(did):
    conn = get_db()
    # Cek apakah ada faktur
    count = conn.execute("SELECT COUNT(*) FROM purchase_invoices WHERE distributor_id=?", (did,)).fetchone()[0]
    if count > 0:
        conn.close()
        return err('Tidak dapat menghapus distributor yang memiliki faktur pembelian')
    conn.execute("UPDATE distributors SET active=0 WHERE id=?", (did,))
    conn.commit()
    conn.close()
    return ok()

@app.route('/api/distributors/<int:did>/invoices', methods=['GET'])
def get_distributor_invoices(did):
    conn = get_db()
    invoices = rows2list(conn.execute(
        "SELECT * FROM purchase_invoices WHERE distributor_id=? ORDER BY invoice_date DESC", (did,)
    ).fetchall())
    conn.close()
    return ok(invoices)

@app.route('/api/distributors/<int:did>/ledger', methods=['GET'])
def get_distributor_ledger(did):
    """Kartu hutang distributor - tampilan buku besar."""
    conn = get_db()
    dist = row2dict(conn.execute("SELECT * FROM distributors WHERE id=?", (did,)).fetchone())
    if not dist:
        conn.close()
        return err('Distributor tidak ditemukan', 404)

    invoices = rows2list(conn.execute(
        "SELECT * FROM purchase_invoices WHERE distributor_id=? ORDER BY invoice_date ASC", (did,)
    ).fetchall())

    ledger = []
    running_balance = 0

    for inv in invoices:
        # Debit: hutang bertambah
        running_balance += inv['total_amount']
        ledger.append({
            'date': inv['invoice_date'],
            'type': 'invoice',
            'reference': inv['invoice_no'],
            'description': f"Faktur Pembelian ({inv['payment_type']})",
            'debit': inv['total_amount'],
            'credit': 0,
            'balance': running_balance
        })
        # Kredit: pembayaran
        payments = rows2list(conn.execute(
            "SELECT * FROM supplier_debt_payments WHERE invoice_id=? ORDER BY payment_date ASC", (inv['id'],)
        ).fetchall())
        for pay in payments:
            running_balance -= pay['amount']
            ledger.append({
                'date': pay['payment_date'],
                'type': 'payment',
                'reference': f"PAY-{pay['id']:04d}",
                'description': f"Pembayaran Hutang ({pay['method']})" + (f" - {pay['notes']}" if pay['notes'] else ''),
                'debit': 0,
                'credit': pay['amount'],
                'balance': running_balance
            })

    summary = row2dict(conn.execute("""
        SELECT COALESCE(SUM(total_amount),0) as total_purchase,
               COALESCE(SUM(paid_amount),0) as total_paid,
               COALESCE(SUM(remaining_amount),0) as total_debt,
               COUNT(*) as total_invoices,
               COUNT(CASE WHEN status IN ('unpaid','partial','overdue') THEN 1 END) as active_invoices,
               COUNT(CASE WHEN status='overdue' THEN 1 END) as overdue_invoices
        FROM purchase_invoices WHERE distributor_id=?
    """, (did,)).fetchone())

    conn.close()
    return ok({
        'distributor': dist,
        'ledger': ledger,
        'summary': summary
    })

# ─── PEMBELIAN BARANG ────────────────────────────────────────────────────────

@app.route('/api/purchases', methods=['GET'])
def get_purchases():
    recalculate_all_invoice_statuses()
    conn = get_db()
    limit  = int(request.args.get('limit', 100))
    offset = int(request.args.get('offset', 0))
    status = request.args.get('status')
    date_from = request.args.get('from')
    date_to   = request.args.get('to')
    dist_id   = request.args.get('distributor_id')
    search    = request.args.get('search','')

    q = "SELECT * FROM purchase_invoices WHERE 1=1"
    params = []
    if status and status != 'all': q += " AND status=?"; params.append(status)
    if date_from: q += " AND invoice_date >= ?"; params.append(date_from)
    if date_to:   q += " AND invoice_date <= ?"; params.append(date_to)
    if dist_id:   q += " AND distributor_id=?"; params.append(int(dist_id))
    if search:    q += " AND (invoice_no LIKE ? OR distributor_name LIKE ?)"; params += [f'%{search}%', f'%{search}%']
    q += " ORDER BY invoice_date DESC, id DESC LIMIT ? OFFSET ?"
    params += [limit, offset]

    invoices = rows2list(conn.execute(q, params).fetchall())
    total = conn.execute("SELECT COUNT(*) FROM purchase_invoices WHERE 1=1" +
                         (" AND status=?" if status and status != 'all' else ""),
                         ([status] if status and status != 'all' else [])).fetchone()[0]
    conn.close()
    return ok(invoices, total=total)

@app.route('/api/purchases', methods=['POST'])
def add_purchase():
    d = request.json or {}
    items = d.get('items', [])
    if not items: return err('Item pembelian tidak boleh kosong')
    if not d.get('distributor_id'): return err('Distributor wajib dipilih')
    if not d.get('invoice_no'): return err('Nomor faktur wajib diisi')
    if not d.get('invoice_date'): return err('Tanggal faktur wajib diisi')

    payment_type = d.get('payment_type', 'cash')
    if payment_type not in ('cash','hutang_penuh','hutang_sebagian'):
        return err('Tipe pembayaran tidak valid')

    conn = get_db()
    try:
        # Cek duplikat nomor faktur
        existing = conn.execute("SELECT id FROM purchase_invoices WHERE invoice_no=?", (d['invoice_no'],)).fetchone()
        if existing:
            conn.close()
            return err(f"Nomor faktur {d['invoice_no']} sudah ada")

        dist = row2dict(conn.execute("SELECT * FROM distributors WHERE id=?", (d['distributor_id'],)).fetchone())
        if not dist:
            conn.close()
            return err('Distributor tidak ditemukan')

        # Hitung total
        total_amount = sum(float(i.get('qty',0)) * float(i.get('buy_price',0)) for i in items)
        if total_amount <= 0:
            conn.close()
            return err('Total pembelian tidak valid')

        # Tentukan paid_amount dan remaining
        dp = float(d.get('down_payment', 0))
        if payment_type == 'cash':
            paid_amount = total_amount
            remaining   = 0
            status      = 'paid'
        elif payment_type == 'hutang_penuh':
            paid_amount = 0
            remaining   = total_amount
            status      = 'unpaid'
        else:  # hutang_sebagian
            paid_amount = min(dp, total_amount)
            remaining   = total_amount - paid_amount
            status      = 'partial' if paid_amount > 0 else 'unpaid'

        user_id   = d.get('user_id')
        user_name = d.get('user_name', 'System')

        # Insert faktur
        c = conn.execute("""INSERT INTO purchase_invoices
            (invoice_no,invoice_date,due_date,distributor_id,distributor_name,payment_type,
             total_amount,paid_amount,remaining_amount,status,notes,
             created_by,created_by_name,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d['invoice_no'], d['invoice_date'], d.get('due_date'),
             dist['id'], dist['name'], payment_type,
             total_amount, paid_amount, remaining, status,
             d.get('notes',''), user_id, user_name, now_iso(), now_iso()))
        inv_id = c.lastrowid

        # Insert items + update stok
        for item in items:
            pid   = item.get('product_id')
            qty   = float(item.get('qty', 0))
            price = float(item.get('buy_price', 0))
            sub   = qty * price
            prod_name = item.get('product_name', '')

            if pid:
                prod = row2dict(conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone())
                if prod:
                    prod_name = prod['name']
                    before = prod['stock']
                    after  = before + qty
                    conn.execute("UPDATE products SET stock=?, buy_price=? WHERE id=?",
                                 (after, price, pid))
                    conn.execute("""INSERT INTO stock_log
                                    (product_id,product_name,delta,reason,before,after,user_id,date)
                                    VALUES (?,?,?,?,?,?,?,?)""",
                                 (pid, prod_name, qty, f"Pembelian {d['invoice_no']}",
                                  before, after, user_id, now_iso()))

            conn.execute("""INSERT INTO purchase_invoice_items
                            (invoice_id,product_id,product_name,qty,buy_price,subtotal)
                            VALUES (?,?,?,?,?,?)""",
                         (inv_id, pid, prod_name, qty, price, sub))

        # Kas otomatis
        if payment_type == 'cash':
            add_cash_transaction(conn, 'out', 'Pembelian Barang Cash',
                                 d['invoice_no'], f"Pembelian dari {dist['name']}",
                                 total_amount, 'purchase_invoice', inv_id,
                                 user_id, user_name, d['invoice_date'])
        elif payment_type == 'hutang_sebagian' and paid_amount > 0:
            add_cash_transaction(conn, 'out', 'DP Pembelian Barang',
                                 d['invoice_no'], f"DP Pembelian dari {dist['name']}",
                                 paid_amount, 'purchase_invoice', inv_id,
                                 user_id, user_name, d['invoice_date'])
            # Catat payment record juga
            conn.execute("""INSERT INTO supplier_debt_payments
                            (invoice_id,payment_date,amount,method,notes,created_by,created_by_name,created_at)
                            VALUES (?,?,?,?,?,?,?,?)""",
                         (inv_id, d['invoice_date'], paid_amount, d.get('payment_method','cash'),
                          'DP saat barang datang', user_id, user_name, now_iso()))

        conn.commit()
        inv = row2dict(conn.execute("SELECT * FROM purchase_invoices WHERE id=?", (inv_id,)).fetchone())
        conn.close()
        log_activity(user_id, 'purchase', f"Faktur {d['invoice_no']} Rp {total_amount:,.0f} dari {dist['name']}")
        return ok(inv)

    except Exception as e:
        conn.rollback()
        conn.close()
        traceback.print_exc()
        return err(str(e), 500)

@app.route('/api/purchases/<int:inv_id>', methods=['GET'])
def get_purchase_detail(inv_id):
    conn = get_db()
    inv = row2dict(conn.execute("SELECT * FROM purchase_invoices WHERE id=?", (inv_id,)).fetchone())
    if not inv:
        conn.close()
        return err('Faktur tidak ditemukan', 404)
    items = rows2list(conn.execute(
        "SELECT * FROM purchase_invoice_items WHERE invoice_id=?", (inv_id,)
    ).fetchall())
    payments = rows2list(conn.execute(
        "SELECT * FROM supplier_debt_payments WHERE invoice_id=? ORDER BY payment_date ASC", (inv_id,)
    ).fetchall())
    conn.close()
    return ok({'invoice': inv, 'items': items, 'payments': payments})

@app.route('/api/purchases/next-invoice-no', methods=['GET'])
def next_invoice_no():
    conn = get_db()
    today = datetime.date.today()
    prefix = f"PO/{today.strftime('%Y%m')}"
    last = conn.execute(
        "SELECT invoice_no FROM purchase_invoices WHERE invoice_no LIKE ? ORDER BY id DESC LIMIT 1",
        (f"{prefix}%",)
    ).fetchone()
    conn.close()
    if last:
        try:
            num = int(last['invoice_no'].split('/')[-1]) + 1
        except:
            num = 1
    else:
        num = 1
    return ok(f"{prefix}/{num:04d}")

# ─── HUTANG DISTRIBUTOR ───────────────────────────────────────────────────────

@app.route('/api/supplier-debts', methods=['GET'])
def get_supplier_debts():
    recalculate_all_invoice_statuses()
    conn = get_db()
    status = request.args.get('status', 'all')
    date_from = request.args.get('from')
    date_to   = request.args.get('to')
    search    = request.args.get('search','')

    q = """SELECT pi.*, d.sales_name, d.phone as dist_phone
           FROM purchase_invoices pi
           LEFT JOIN distributors d ON pi.distributor_id = d.id
           WHERE pi.payment_type != 'cash'"""
    params = []
    if status and status != 'all':
        q += " AND pi.status=?"; params.append(status)
    if date_from: q += " AND pi.invoice_date >= ?"; params.append(date_from)
    if date_to:   q += " AND pi.invoice_date <= ?"; params.append(date_to)
    if search:
        q += " AND (pi.invoice_no LIKE ? OR pi.distributor_name LIKE ?)"
        params += [f'%{search}%', f'%{search}%']
    q += " ORDER BY pi.due_date ASC, pi.invoice_date DESC"

    debts = rows2list(conn.execute(q, params).fetchall())
    summary = row2dict(conn.execute("""
        SELECT COALESCE(SUM(remaining_amount),0) as total_debt,
               COUNT(CASE WHEN status='overdue' THEN 1 END) as overdue_count,
               COUNT(CASE WHEN status IN ('unpaid','partial') THEN 1 END) as active_count,
               COALESCE(SUM(CASE WHEN status='overdue' THEN remaining_amount ELSE 0 END),0) as overdue_amount
        FROM purchase_invoices WHERE payment_type != 'cash'
    """).fetchone())
    conn.close()
    return ok(debts, summary=summary)

@app.route('/api/supplier-debts/<int:inv_id>/pay', methods=['POST'])
def pay_supplier_debt(inv_id):
    d = request.json or {}
    amount = float(d.get('amount', 0))
    if amount <= 0: return err('Nominal pembayaran harus lebih dari 0')

    conn = get_db()
    try:
        inv = row2dict(conn.execute("SELECT * FROM purchase_invoices WHERE id=?", (inv_id,)).fetchone())
        if not inv:
            conn.close()
            return err('Faktur tidak ditemukan', 404)
        if inv['remaining_amount'] <= 0:
            conn.close()
            return err('Faktur sudah lunas')
        if amount > inv['remaining_amount'] + 0.01:
            conn.close()
            return err(f"Pembayaran melebihi sisa hutang (Rp {inv['remaining_amount']:,.0f})")

        user_id   = d.get('user_id')
        user_name = d.get('user_name', 'System')
        pay_date  = d.get('payment_date', today_str())
        method    = d.get('method', 'transfer')
        notes     = d.get('notes', '')

        # Insert payment
        conn.execute("""INSERT INTO supplier_debt_payments
                        (invoice_id,payment_date,amount,method,notes,created_by,created_by_name,created_at)
                        VALUES (?,?,?,?,?,?,?,?)""",
                     (inv_id, pay_date, amount, method, notes, user_id, user_name, now_iso()))

        # Update invoice status
        update_invoice_status(conn, inv_id)

        # Kas keluar
        add_cash_transaction(conn, 'out', 'Pembayaran Hutang Distributor',
                             inv['invoice_no'], f"Bayar hutang {inv['distributor_name']}",
                             amount, 'supplier_debt_payment', inv_id,
                             user_id, user_name, pay_date)

        conn.commit()
        inv_updated = row2dict(conn.execute("SELECT * FROM purchase_invoices WHERE id=?", (inv_id,)).fetchone())
        conn.close()
        log_activity(user_id, 'pay_supplier', f"Bayar hutang {inv['invoice_no']}: Rp {amount:,.0f}")
        return ok(inv_updated)

    except Exception as e:
        conn.rollback()
        conn.close()
        traceback.print_exc()
        return err(str(e), 500)

@app.route('/api/supplier-debts/<int:inv_id>/payments', methods=['GET'])
def get_debt_payment_history(inv_id):
    conn = get_db()
    payments = rows2list(conn.execute(
        "SELECT * FROM supplier_debt_payments WHERE invoice_id=? ORDER BY payment_date ASC", (inv_id,)
    ).fetchall())
    conn.close()
    return ok(payments)

# ─── KAS ─────────────────────────────────────────────────────────────────────

@app.route('/api/cash', methods=['GET'])
def get_cash_transactions():
    conn = get_db()
    type_  = request.args.get('type')
    date_from = request.args.get('from')
    date_to   = request.args.get('to')
    search    = request.args.get('search','')
    limit  = int(request.args.get('limit', 200))
    offset = int(request.args.get('offset', 0))

    q = "SELECT * FROM cash_transactions WHERE 1=1"
    params = []
    if type_ and type_ != 'all': q += " AND type=?"; params.append(type_)
    if date_from: q += " AND date >= ?"; params.append(date_from)
    if date_to:   q += " AND date <= ?"; params.append(date_to)
    if search:
        q += " AND (category LIKE ? OR reference LIKE ? OR description LIKE ?)"
        params += [f'%{search}%', f'%{search}%', f'%{search}%']
    q += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params += [limit, offset]

    txs = rows2list(conn.execute(q, params).fetchall())
    balance = get_cash_balance(conn)
    masuk  = conn.execute("SELECT COALESCE(SUM(amount),0) FROM cash_transactions WHERE type='in'").fetchone()[0]
    keluar = conn.execute("SELECT COALESCE(SUM(amount),0) FROM cash_transactions WHERE type='out'").fetchone()[0]
    conn.close()
    return ok(txs, balance=balance, total_in=masuk, total_out=keluar)

@app.route('/api/cash', methods=['POST'])
def add_manual_cash():
    """Tambah kas manual (kas masuk/keluar non-otomatis)."""
    d = request.json or {}
    type_  = d.get('type')
    amount = float(d.get('amount', 0))
    if type_ not in ('in','out'): return err('Tipe kas tidak valid')
    if amount <= 0: return err('Nominal harus lebih dari 0')
    if not d.get('category'): return err('Kategori wajib diisi')

    conn = get_db()
    user_name = d.get('user_name','')
    add_cash_transaction(conn, type_, d['category'], d.get('reference',''),
                         d.get('description',''), amount, 'manual', None,
                         d.get('user_id'), user_name, d.get('date', today_str()))
    conn.commit()
    conn.close()
    return ok()

@app.route('/api/cash/balance', methods=['GET'])
def get_balance():
    balance = get_cash_balance()
    return ok({'balance': balance})

# ─── DASHBOARD ───────────────────────────────────────────────────────────────

@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    recalculate_all_invoice_statuses()
    conn = get_db()
    today = today_str()
    month_start = today[:7] + '-01'
    next7 = (datetime.date.today() + datetime.timedelta(days=7)).isoformat()

    # Sales bulan ini
    sales_month = row2dict(conn.execute("""
        SELECT COALESCE(SUM(total),0) as revenue, COUNT(*) as tx_count
        FROM transactions WHERE status='done' AND date >= ? AND date <= ?
    """, (month_start, today + 'T23:59:59')).fetchone())

    # Produk stok rendah
    low_stock = conn.execute(
        "SELECT COUNT(*) FROM products WHERE active=1 AND stock <= min_stock AND stock > 0"
    ).fetchone()[0]
    out_stock = conn.execute(
        "SELECT COUNT(*) FROM products WHERE active=1 AND stock <= 0"
    ).fetchone()[0]

    # Hutang distributor
    debt_stats = row2dict(conn.execute("""
        SELECT COALESCE(SUM(remaining_amount),0) as total_debt,
               COUNT(CASE WHEN status='overdue' THEN 1 END) as overdue_count,
               COALESCE(SUM(CASE WHEN status='overdue' THEN remaining_amount ELSE 0 END),0) as overdue_amount,
               COUNT(CASE WHEN status IN ('unpaid','partial') THEN 1 END) as active_count
        FROM purchase_invoices WHERE payment_type != 'cash'
    """).fetchone())

    # Faktur jatuh tempo hari ini
    due_today = rows2list(conn.execute("""
        SELECT pi.*, d.phone as dist_phone FROM purchase_invoices pi
        LEFT JOIN distributors d ON pi.distributor_id=d.id
        WHERE pi.status IN ('unpaid','partial','overdue') AND pi.due_date = ?
        ORDER BY pi.remaining_amount DESC
    """, (today,)).fetchall())

    # Faktur jatuh tempo 7 hari ke depan
    due_7days = rows2list(conn.execute("""
        SELECT pi.*, d.phone as dist_phone FROM purchase_invoices pi
        LEFT JOIN distributors d ON pi.distributor_id=d.id
        WHERE pi.status IN ('unpaid','partial') AND pi.due_date > ? AND pi.due_date <= ?
        ORDER BY pi.due_date ASC
    """, (today, next7)).fetchall())

    # Faktur overdue (sudah lewat)
    overdue = rows2list(conn.execute("""
        SELECT pi.*, d.phone as dist_phone FROM purchase_invoices pi
        LEFT JOIN distributors d ON pi.distributor_id=d.id
        WHERE pi.status='overdue'
        ORDER BY pi.due_date ASC
    """).fetchall())

    # Pembelian bulan ini
    purchase_month = row2dict(conn.execute("""
        SELECT COALESCE(SUM(total_amount),0) as total, COUNT(*) as count
        FROM purchase_invoices WHERE invoice_date >= ? AND invoice_date <= ?
    """, (month_start, today)).fetchone())

    # Kas
    balance = get_cash_balance(conn)
    kas_masuk_month  = conn.execute("""
        SELECT COALESCE(SUM(amount),0) FROM cash_transactions
        WHERE type='in' AND date >= ? AND date <= ?
    """, (month_start, today)).fetchone()[0]
    kas_keluar_month = conn.execute("""
        SELECT COALESCE(SUM(amount),0) FROM cash_transactions
        WHERE type='out' AND date >= ? AND date <= ?
    """, (month_start, today)).fetchone()[0]

    # Distributor count
    dist_count = conn.execute("SELECT COUNT(*) FROM distributors WHERE active=1").fetchone()[0]

    # Top 10 distributor by hutang
    top_debtors = rows2list(conn.execute("""
        SELECT d.name, d.sales_name,
               COALESCE(SUM(pi.remaining_amount),0) as total_debt,
               COUNT(pi.id) as invoice_count
        FROM distributors d
        LEFT JOIN purchase_invoices pi ON d.id=pi.distributor_id
            AND pi.status IN ('unpaid','partial','overdue')
        GROUP BY d.id
        HAVING total_debt > 0
        ORDER BY total_debt DESC LIMIT 10
    """).fetchall())

    conn.close()
    return ok({
        'sales_month':      sales_month,
        'purchase_month':   purchase_month,
        'debt_stats':       debt_stats,
        'due_today':        due_today,
        'due_7days':        due_7days,
        'overdue':          overdue,
        'cash_balance':     balance,
        'kas_masuk_month':  kas_masuk_month,
        'kas_keluar_month': kas_keluar_month,
        'dist_count':       dist_count,
        'top_debtors':      top_debtors,
        'low_stock':        low_stock,
        'out_stock':        out_stock,
    })

# ─── REPORTS ─────────────────────────────────────────────────────────────────

@app.route('/api/reports/summary', methods=['GET'])
def report_summary():
    date_from = request.args.get('from', today_str())
    date_to   = request.args.get('to',   today_str())
    conn = get_db()
    txs = rows2list(conn.execute("""
        SELECT * FROM transactions
        WHERE status != 'void' AND date >= ? AND date <= ?
    """, (date_from, date_to + 'T23:59:59')).fetchall())
    tx_ids = [t['id'] for t in txs]
    items = rows2list(conn.execute(
        "SELECT * FROM transaction_items WHERE transaction_id IN ({})".format(
            ','.join(['?']*len(tx_ids)) if tx_ids else ['NULL']
        ), tx_ids if tx_ids else []
    ).fetchall()) if tx_ids else []
    products = rows2list(conn.execute("SELECT * FROM products WHERE active=1").fetchall())
    debts    = rows2list(conn.execute("SELECT * FROM debts").fetchall())
    prod_map = {p['id']: p for p in products}
    revenue  = sum(t['total'] for t in txs)
    tx_count = len(txs)
    items_sold = sum(i['qty'] for i in items)
    cogs = sum((prod_map.get(i['product_id'], {}).get('buy_price') or i.get('buy_price') or 0) * i['qty'] for i in items)
    profit = revenue - cogs - sum(t['discount'] for t in txs)
    method_break = {}
    for t in txs:
        method_break[t['payment_method']] = method_break.get(t['payment_method'],0) + t['total']
    prod_sales = {}
    for i in items:
        pid = i['product_id']
        if pid not in prod_sales:
            prod_sales[pid] = {'name': i['name'], 'qty': 0, 'revenue': 0}
        prod_sales[pid]['qty']     += i['qty']
        prod_sales[pid]['revenue'] += i['subtotal']
    top_products = sorted(prod_sales.values(), key=lambda x: x['revenue'], reverse=True)[:10]
    stock_value      = sum(p['stock'] * p['buy_price'] for p in products)
    stock_sell_value = sum(p['stock'] * p['price'] for p in products)
    total_debt       = sum(max(0, d['amount'] - d['paid']) for d in debts)
    daily = {}
    for t in txs:
        day = t['date'][:10]
        daily[day] = daily.get(day, 0) + t['total']
    conn.close()
    return ok({
        'revenue': revenue, 'profit': profit, 'tx_count': tx_count,
        'items_sold': items_sold, 'method_break': method_break,
        'top_products': top_products, 'stock_value': stock_value,
        'stock_sell_value': stock_sell_value, 'total_debt': total_debt,
        'daily': [{'date': k, 'total': v} for k, v in sorted(daily.items())],
    })

@app.route('/api/reports/cash-flow', methods=['GET'])
def report_cash_flow():
    date_from = request.args.get('from', today_str()[:7] + '-01')
    date_to   = request.args.get('to', today_str())
    conn = get_db()
    txs = rows2list(conn.execute("""
        SELECT * FROM cash_transactions WHERE date >= ? AND date <= ?
        ORDER BY date ASC, id ASC
    """, (date_from, date_to)).fetchall())

    masuk  = sum(t['amount'] for t in txs if t['type'] == 'in')
    keluar = sum(t['amount'] for t in txs if t['type'] == 'out')

    by_category_in  = {}
    by_category_out = {}
    for t in txs:
        cat = t['category'] or 'Lainnya'
        if t['type'] == 'in':
            by_category_in[cat] = by_category_in.get(cat, 0) + t['amount']
        else:
            by_category_out[cat] = by_category_out.get(cat, 0) + t['amount']

    conn.close()
    return ok({
        'transactions': txs,
        'total_in': masuk,
        'total_out': keluar,
        'net': masuk - keluar,
        'by_category_in': by_category_in,
        'by_category_out': by_category_out,
    })

@app.route('/api/reports/purchases', methods=['GET'])
def report_purchases():
    recalculate_all_invoice_statuses()
    date_from = request.args.get('from', today_str()[:7] + '-01')
    date_to   = request.args.get('to', today_str())
    conn = get_db()
    invoices = rows2list(conn.execute("""
        SELECT * FROM purchase_invoices WHERE invoice_date >= ? AND invoice_date <= ?
        ORDER BY invoice_date DESC
    """, (date_from, date_to)).fetchall())
    total = sum(i['total_amount'] for i in invoices)
    total_paid = sum(i['paid_amount'] for i in invoices)
    total_debt = sum(i['remaining_amount'] for i in invoices)
    conn.close()
    return ok({'invoices': invoices, 'total': total, 'total_paid': total_paid, 'total_debt': total_debt})

@app.route('/api/reports/debts', methods=['GET'])
def report_debts():
    recalculate_all_invoice_statuses()
    conn = get_db()
    status = request.args.get('status', 'all')
    q = """SELECT pi.*, d.sales_name, d.phone as dist_phone
           FROM purchase_invoices pi
           LEFT JOIN distributors d ON pi.distributor_id=d.id
           WHERE pi.payment_type != 'cash'"""
    params = []
    if status and status != 'all':
        q += " AND pi.status=?"; params.append(status)
    q += " ORDER BY pi.due_date ASC"
    debts = rows2list(conn.execute(q, params).fetchall())
    summary = row2dict(conn.execute("""
        SELECT COALESCE(SUM(remaining_amount),0) as total_debt,
               COALESCE(SUM(total_amount),0) as total_purchase,
               COALESCE(SUM(paid_amount),0) as total_paid
        FROM purchase_invoices WHERE payment_type != 'cash'
    """).fetchone())
    conn.close()
    return ok({'debts': debts, 'summary': summary})

# ─── EXPORT EXCEL ────────────────────────────────────────────────────────────

@app.route('/api/export/purchases', methods=['GET'])
def export_purchases_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return err('openpyxl tidak terinstall. Jalankan: pip install openpyxl')

    date_from = request.args.get('from', today_str()[:7] + '-01')
    date_to   = request.args.get('to', today_str())
    recalculate_all_invoice_statuses()
    conn = get_db()
    invoices = rows2list(conn.execute("""
        SELECT * FROM purchase_invoices WHERE invoice_date >= ? AND invoice_date <= ?
        ORDER BY invoice_date ASC
    """, (date_from, date_to)).fetchall())
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Pembelian"

    hdr_fill = PatternFill("solid", fgColor="166534")
    hdr_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="E8F5EC")
    total_font = Font(bold=True)

    headers = ['No','No Faktur','Tgl Faktur','Jatuh Tempo','Distributor',
               'Tipe Bayar','Total','Dibayar','Sisa','Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    for i, inv in enumerate(invoices, 1):
        status_map = {'paid':'Lunas','unpaid':'Belum Lunas','partial':'Sebagian','overdue':'Jatuh Tempo'}
        type_map   = {'cash':'Tunai','hutang_penuh':'Hutang Penuh','hutang_sebagian':'Hutang Sebagian'}
        ws.append([
            i, inv['invoice_no'], inv['invoice_date'], inv['due_date'] or '-',
            inv['distributor_name'], type_map.get(inv['payment_type'], inv['payment_type']),
            inv['total_amount'], inv['paid_amount'], inv['remaining_amount'],
            status_map.get(inv['status'], inv['status'])
        ])

    # Total row
    total_row = ws.max_row + 1
    ws.cell(total_row, 1, 'TOTAL')
    ws.cell(total_row, 7, sum(i['total_amount'] for i in invoices))
    ws.cell(total_row, 8, sum(i['paid_amount'] for i in invoices))
    ws.cell(total_row, 9, sum(i['remaining_amount'] for i in invoices))
    for col in range(1, 11):
        ws.cell(total_row, col).font = total_font
        ws.cell(total_row, col).fill = total_fill

    # Format currency columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in [7, 8, 9]:
            row[col_idx-1].number_format = '#,##0'

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"laporan-pembelian-{date_from}-{date_to}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

@app.route('/api/export/debts', methods=['GET'])
def export_debts_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return err('openpyxl tidak terinstall')

    recalculate_all_invoice_statuses()
    conn = get_db()
    debts = rows2list(conn.execute("""
        SELECT pi.*, d.sales_name, d.phone as dist_phone
        FROM purchase_invoices pi
        LEFT JOIN distributors d ON pi.distributor_id=d.id
        WHERE pi.payment_type != 'cash' AND pi.status != 'paid'
        ORDER BY pi.due_date ASC
    """).fetchall())
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hutang Distributor"
    hdr_fill = PatternFill("solid", fgColor="166534")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ['No','No Faktur','Distributor','Sales','Tgl Faktur','Jatuh Tempo','Total','Dibayar','Sisa','Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    status_map = {'unpaid':'Belum Lunas','partial':'Sebagian','overdue':'JATUH TEMPO','paid':'Lunas'}
    for i, d in enumerate(debts, 1):
        ws.append([i, d['invoice_no'], d['distributor_name'], d.get('sales_name',''),
                   d['invoice_date'], d['due_date'] or '-',
                   d['total_amount'], d['paid_amount'], d['remaining_amount'],
                   status_map.get(d['status'], d['status'])])

    for row in ws.iter_rows(min_row=2):
        for col_idx in [7, 8, 9]:
            row[col_idx-1].number_format = '#,##0'

    for col, width in zip('ABCDEFGHIJ', [5,20,28,16,14,14,16,16,16,14]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"hutang-distributor-{today_str()}.xlsx")

@app.route('/api/export/cash', methods=['GET'])
def export_cash_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return err('openpyxl tidak terinstall')

    date_from = request.args.get('from', today_str()[:7] + '-01')
    date_to   = request.args.get('to', today_str())
    conn = get_db()
    txs = rows2list(conn.execute("""
        SELECT * FROM cash_transactions WHERE date >= ? AND date <= ? ORDER BY id ASC
    """, (date_from, date_to)).fetchall())
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arus Kas"
    hdr_fill = PatternFill("solid", fgColor="166534")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ['No','Tanggal','Tipe','Kategori','Referensi','Keterangan','Kas Masuk','Kas Keluar','Saldo']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    for i, t in enumerate(txs, 1):
        ws.append([i, t['date'],
                   'Masuk' if t['type']=='in' else 'Keluar',
                   t['category'] or '', t['reference'] or '', t['description'] or '',
                   t['amount'] if t['type']=='in' else 0,
                   t['amount'] if t['type']=='out' else 0,
                   t['balance']])

    for row in ws.iter_rows(min_row=2):
        for col_idx in [7, 8, 9]:
            row[col_idx-1].number_format = '#,##0'

    for col, width in zip('ABCDEFGHI', [5,14,8,24,20,30,16,16,16]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"arus-kas-{date_from}-{date_to}.xlsx")

# ─── USERS ───────────────────────────────────────────────────────────────────

@app.route('/api/users', methods=['GET'])
def get_users():
    conn = get_db()
    rows = rows2list(conn.execute("SELECT id,username,role,name,active FROM users").fetchall())
    conn.close()
    return ok(rows)

@app.route('/api/users', methods=['POST'])
def add_user():
    d = request.json or {}
    if not d.get('username') or not d.get('name') or not d.get('password'):
        return err('Data tidak lengkap')
    conn = get_db()
    if conn.execute("SELECT id FROM users WHERE username=?", (d['username'],)).fetchone():
        conn.close(); return err('Username sudah digunakan')
    hashed = hashlib.sha256(d['password'].encode()).hexdigest()
    c = conn.execute("INSERT INTO users (username,password,role,name) VALUES (?,?,?,?)",
                     (d['username'], hashed, d.get('role','cashier'), d['name']))
    conn.commit()
    uid = c.lastrowid
    conn.close()
    return ok({'id': uid, 'username': d['username'], 'role': d.get('role','cashier'), 'name': d['name']})

@app.route('/api/users/<int:uid>', methods=['PUT'])
def update_user(uid):
    d = request.json or {}
    conn = get_db()
    if d.get('password'):
        hashed = hashlib.sha256(d['password'].encode()).hexdigest()
        conn.execute("UPDATE users SET name=?,role=?,password=? WHERE id=?", (d.get('name',''), d.get('role','cashier'), hashed, uid))
    else:
        conn.execute("UPDATE users SET name=?,role=? WHERE id=?", (d.get('name',''), d.get('role','cashier'), uid))
    conn.commit()
    conn.close()
    return ok()

@app.route('/api/activity_log', methods=['GET'])
def get_activity_log():
    limit = int(request.args.get('limit', 100))
    conn = get_db()
    rows = rows2list(conn.execute("SELECT * FROM activity_log ORDER BY id DESC LIMIT ?", (limit,)).fetchall())
    conn.close()
    return ok(rows)

@app.route('/api/export/price-labels', methods=['POST'])
def export_price_labels():
    """
    Export label harga produk ke PDF format A4.
    Layout: 4 kolom × 10 baris = 40 label per halaman.
    Body JSON: { product_ids: [1,2,3,...] }
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor, black, white
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        return err('reportlab tidak terinstall. Jalankan: pip install reportlab')

    d = request.json or {}
    product_ids = d.get('product_ids', [])
    if not product_ids:
        return err('Pilih minimal 1 produk')

    conn = get_db()
    settings_rows = conn.execute("SELECT key,value FROM settings").fetchall()
    settings = {r['key']: r['value'] for r in settings_rows}
    store_name = settings.get('storeName', 'Toko Kelontong')

    placeholders = ','.join(['?'] * len(product_ids))
    products = rows2list(conn.execute(
        f"SELECT * FROM products WHERE id IN ({placeholders}) AND active=1 ORDER BY category, name",
        product_ids
    ).fetchall())
    conn.close()

    if not products:
        return err('Produk tidak ditemukan')

    buf = io.BytesIO()

    # ── Ukuran halaman & margin ──
    PAGE_W, PAGE_H = A4           # 210 × 297 mm
    MARGIN_L  = 5  * mm
    MARGIN_R  = 5  * mm
    MARGIN_T  = 8  * mm
    MARGIN_B  = 8  * mm
    COLS      = 4
    ROWS      = 10
    PER_PAGE  = COLS * ROWS       # 40

    LABEL_W = (PAGE_W - MARGIN_L - MARGIN_R) / COLS   # ~50mm
    LABEL_H = (PAGE_H - MARGIN_T - MARGIN_B) / ROWS   # ~28.1mm

    GAP = 0.6 * mm   # gap antar label (garis potong)

    # Warna
    COL_GREEN      = HexColor('#166534')
    COL_GREEN_LITE = HexColor('#f0faf2')
    COL_GREEN_MID  = HexColor('#dcfce7')
    COL_GRAY       = HexColor('#6b7280')
    COL_DARK       = HexColor('#111827')
    COL_BORDER     = HexColor('#d1e8d5')
    COL_CAT_BG     = HexColor('#166534')

    c = rl_canvas.Canvas(buf, pagesize=A4)

    def draw_label(x, y, prod):
        """Gambar satu label pada posisi (x,y) sudut kiri bawah."""
        w = LABEL_W - GAP
        h = LABEL_H - GAP

        # Background label
        c.setFillColor(white)
        c.setStrokeColor(COL_BORDER)
        c.setLineWidth(0.4)
        c.roundRect(x, y, w, h, 2*mm, fill=1, stroke=1)

        # Strip header — nama toko
        hdr_h = 5.2 * mm
        c.setFillColor(COL_GREEN)
        c.roundRect(x, y + h - hdr_h, w, hdr_h, 1.5*mm, fill=1, stroke=0)
        # Pojok bawah header kotak (override rounded supaya rapi)
        c.rect(x, y + h - hdr_h, w, hdr_h * 0.5, fill=1, stroke=0)

        c.setFillColor(white)
        c.setFont('Helvetica-Bold', 5.5)
        c.drawCentredString(x + w/2, y + h - hdr_h + 1.6*mm, store_name.upper())

        # Strip kategori
        cat_h = 3.8 * mm
        cat_y = y + h - hdr_h - cat_h
        c.setFillColor(COL_GREEN_MID)
        c.rect(x, cat_y, w, cat_h, fill=1, stroke=0)
        c.setFillColor(COL_GREEN)
        c.setFont('Helvetica', 5)
        cat_label = (prod.get('category') or 'Umum').upper()
        c.drawCentredString(x + w/2, cat_y + 1.1*mm, cat_label)

        # Area nama produk
        name_area_y = y + 7.5 * mm
        name_area_h = cat_y - name_area_y
        name = prod['name']

        # Potong nama jika terlalu panjang
        c.setFillColor(COL_DARK)
        # Coba fit dalam 2 baris
        c.setFont('Helvetica-Bold', 6.8)
        max_w = w - 3*mm

        def split_text(text, font, size, max_width):
            words = text.split(' ')
            lines = []
            cur = ''
            for word in words:
                test = (cur + ' ' + word).strip()
                if c.stringWidth(test, font, size) <= max_width:
                    cur = test
                else:
                    if cur:
                        lines.append(cur)
                    cur = word
            if cur:
                lines.append(cur)
            return lines

        lines = split_text(name, 'Helvetica-Bold', 6.8, max_w)
        if len(lines) > 2:
            # Truncate
            lines = lines[:2]
            if len(lines[1]) > 3:
                lines[1] = lines[1][:-3] + '...'

        line_h = 7.2 * mm
        total_text_h = len(lines) * line_h * 0.72
        start_y = name_area_y + name_area_h/2 + total_text_h/2

        for i, line in enumerate(lines):
            c.setFont('Helvetica-Bold', 6.8)
            c.setFillColor(COL_DARK)
            c.drawCentredString(x + w/2, start_y - i * line_h * 0.72, line)

        # Harga — area bawah
        price_area_h = 7 * mm
        price_bg_y = y + 0.5*mm

        c.setFillColor(COL_GREEN_LITE)
        c.roundRect(x + 1.5*mm, price_bg_y, w - 3*mm, price_area_h, 1*mm, fill=1, stroke=0)

        price_str = 'Rp {:,.0f}'.format(prod['price']).replace(',', '.')
        c.setFillColor(COL_GREEN)
        c.setFont('Helvetica-Bold', 9.5)

        # Auto-shrink font jika harga panjang
        font_size = 9.5
        while c.stringWidth(price_str, 'Helvetica-Bold', font_size) > max_w - 2*mm and font_size > 6:
            font_size -= 0.5

        c.setFont('Helvetica-Bold', font_size)
        c.drawCentredString(x + w/2, price_bg_y + 2*mm, price_str)

    # ── Render semua label ──
    total = len(products)
    pages = (total + PER_PAGE - 1) // PER_PAGE

    for page_idx in range(pages):
        if page_idx > 0:
            c.showPage()

        page_prods = products[page_idx * PER_PAGE : (page_idx + 1) * PER_PAGE]

        for i, prod in enumerate(page_prods):
            col = i % COLS
            row = i // COLS

            x = MARGIN_L + col * LABEL_W
            # Baris dari atas ke bawah
            y = PAGE_H - MARGIN_T - (row + 1) * LABEL_H + GAP/2

            draw_label(x, y, prod)

        # Footer halaman
        c.setFont('Helvetica', 5)
        c.setFillColor(COL_GRAY)
        c.drawCentredString(PAGE_W/2, 4*mm, f'TokoKu POS — Label Harga — {today_str()} — Hal. {page_idx+1}/{pages}')

    c.save()
    buf.seek(0)
    fname = f"label-harga-{today_str()}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=False, download_name=fname)

@app.route('/api/backup', methods=['GET'])
def backup():
    conn = get_db()
    data = {
        '_meta': {'version': 2, 'date': now_iso()},
        'products':      rows2list(conn.execute("SELECT * FROM products").fetchall()),
        'categories':    rows2list(conn.execute("SELECT * FROM categories").fetchall()),
        'customers':     rows2list(conn.execute("SELECT * FROM customers").fetchall()),
        'transactions':  rows2list(conn.execute("SELECT * FROM transactions").fetchall()),
        'transaction_items': rows2list(conn.execute("SELECT * FROM transaction_items").fetchall()),
        'debts':         rows2list(conn.execute("SELECT * FROM debts").fetchall()),
        'debt_payments': rows2list(conn.execute("SELECT * FROM debt_payments").fetchall()),
        'distributors':  rows2list(conn.execute("SELECT * FROM distributors").fetchall()),
        'purchase_invoices': rows2list(conn.execute("SELECT * FROM purchase_invoices").fetchall()),
        'purchase_invoice_items': rows2list(conn.execute("SELECT * FROM purchase_invoice_items").fetchall()),
        'supplier_debt_payments': rows2list(conn.execute("SELECT * FROM supplier_debt_payments").fetchall()),
        'cash_transactions': rows2list(conn.execute("SELECT * FROM cash_transactions").fetchall()),
        'settings':      rows2list(conn.execute("SELECT * FROM settings").fetchall()),
    }
    conn.close()
    buf = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
    buf.seek(0)
    fname = f"tokoku-backup-{datetime.date.today().isoformat()}.json"
    return send_file(buf, mimetype='application/json', as_attachment=True, download_name=fname)

# ─── FRONTEND ────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('static', path)

# ─── MAIN ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    print("="*55)
    print("  TokoKu POS Server v2.0")
    print("  http://0.0.0.0:80")
    print("  Modul: Distributor, Pembelian, Hutang, Kas, Laporan")
    print("="*55)
    import sys
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 80
    try:
        app.run(host='0.0.0.0', port=port, debug=False)
    except PermissionError:
        print(f"\n[!] Port {port} butuh admin. Jalankan: python app.py 8080")
        sys.exit(1)
